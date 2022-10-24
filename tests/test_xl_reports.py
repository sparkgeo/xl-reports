import os
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import xl_reports


@pytest.fixture
def sample_report_template():
    return Path("tests/sample.xlsx")


@pytest.fixture
def report_config():
    return [
        {"cell": "B1", "data_key": "account", "sheet": "my_sheet"},
        {"cell": "B2", "data_key": "report_date", "sheet": "my_sheet"},
        {"range": "A5", "data_key": "report_data", "sheet": "my_sheet"},
    ]


@pytest.fixture
def report_data():
    return {
        "account": "Engineering",
        "report_date": str(date.today()),
        "report_data": [
            [23.43, 11.96, 0.43],
            [6.99, 65.87, 0.6545],
        ],
    }


@pytest.fixture
def report_path():
    pth = "tests/test.xlsx"
    yield pth
    if Path(pth).exists():
        os.remove(pth)


def test_report_generation(
    sample_report_template, report_config, report_data, report_path
):
    xl_reports.generate(report_path, sample_report_template, report_config, report_data)

    wb = load_workbook(report_path)
    ws = wb.worksheets[0]

    assert (
        ws[report_config[0]["cell"]].value == report_data[report_config[0]["data_key"]]
    )
    assert ws["B6"].value == report_data[report_config[2]["data_key"]][1][1]
