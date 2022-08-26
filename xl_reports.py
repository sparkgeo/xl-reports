from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Union, overload

from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.utils.cell import coordinate_to_tuple  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

CellValue = Union[None, str, float, int, datetime]
CellValues = Iterable[Iterable[CellValue]]


def get_worksheet(wb: Workbook, name: str) -> Worksheet:
    if name:
        name = name.lower()
        return next((ws for ws in wb.worksheets if ws.title.lower() == name), None)
    return wb.worksheets[0]


def insert_cell(ws: Worksheet, cell_location: str, value: CellValue) -> None:
    ws[cell_location].value = value


def insert_range(
    ws: Worksheet,
    rng: str,
    data: CellValues,
) -> None:
    _data = data or []
    row, col = coordinate_to_tuple(rng)
    for n, row_val in enumerate(_data):
        for m, col_val in enumerate(row_val):
            ws.cell(row=row + n, column=col + m, value=col_val)


@overload
def insert_data(ws: Worksheet, item: dict, data: CellValue) -> None:
    ...


@overload
def insert_data(ws: Worksheet, item: dict, data: CellValues) -> None:
    ...


def insert_data(ws, item, data) -> None:
    if "cell" in item:
        insert_cell(ws, item["cell"], data[item["data_key"]])
    elif "range" in item:
        insert_range(ws, item["range"], data[item["data_key"]])


def generate(
    report_path: Union[str, Path],
    template: Union[str, Path],
    config: dict,
    data: Dict[
        str,
        Union[CellValue, CellValues],
    ],
):
    wb = load_workbook(template)
    for item in config:
        ws = get_worksheet(wb, item.get("sheet"))
        insert_data(ws, item, data)

    wb.save(report_path)
